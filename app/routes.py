# 1-10: Importing modules
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app
import os  # OS module for file operations
from app.ocr import extract_text_from_image, extract_data  # Import OCR functions
from app.utils import save_uploaded_file, save_to_excel, cleanup_temp_files, validate_extracted_data, allowed_file  # Import utility functions

# 11-20: Blueprint creation
main_bp = Blueprint('main', __name__)  # Create main blueprint for routes

@main_bp.route('/')
def index():
    """
    21-30: Home page route - render file upload form
    """
    # Render the main upload page
    return render_template('index.html')

@main_bp.route('/upload', methods=['POST'])
def upload_files():
    """
    31-100: Handle file upload and OCR processing with improved validation
    """
    print("🚀 Starting file upload process...")
    
    # 11-20: Validate and read uploaded files
    if 'files' not in request.files:
        flash('No files selected', 'error')
        return redirect(url_for('main.index'))
    
    uploaded_files = request.files.getlist('files')
    
    if not uploaded_files or all(file.filename == '' for file in uploaded_files):
        flash('No files selected', 'error')
        return redirect(url_for('main.index'))
    
    # Initialize processing variables
    processed_data = []
    saved_file_paths = []
    upload_folder = current_app.config['UPLOAD_FOLDER']
    skipped_files = []
    
    os.makedirs(upload_folder, exist_ok=True)
    print(f"📁 Upload folder: {upload_folder}")
    
    # 21-40: Process each uploaded file with format validation
    for file in uploaded_files:
        if file and file.filename != '':
            print(f"\n📄 Processing file: {file.filename}")
            
            # Check file format before processing
            if not allowed_file(file.filename):
                print(f"❌ Unsupported format: {file.filename}")
                skipped_files.append(file.filename)
                continue
            
            try:
                # Save uploaded file with secure filename
                file_path = save_uploaded_file(file, upload_folder)
                
                if file_path:
                    saved_file_paths.append(file_path)
                    print(f"💾 File saved: {file_path}")
                    
                    # 41-60: Perform OCR on supported formats
                    raw_text = extract_text_from_image(file_path)
                    
                    if raw_text:
                        print(f"📝 Extracted text preview: {raw_text[:100]}...")
                        
                        # Extract structured data from raw text
                        structured_data = extract_data(raw_text)
                        
                        # Validate extracted data
                        if validate_extracted_data(structured_data):
                            structured_data['filename'] = file.filename
                            processed_data.append(structured_data)
                            print(f"✅ Data extracted for: {file.filename}")
                        else:
                            print(f"⚠️ No valid data extracted from {file.filename}")
                            flash(f'No valid data extracted from {file.filename}', 'warning')
                    else:
                        print(f"❌ No text found in {file.filename}")
                        flash(f'No text found in {file.filename}', 'warning')
                else:
                    flash(f'Failed to save {file.filename}', 'error')
                    
            except Exception as e:
                print(f"❌ Error processing {file.filename}: {str(e)}")
                flash(f'Error processing {file.filename}: {str(e)}', 'error')
                continue
    
    # Handle results and messages
    if skipped_files:
        flash(f'Skipped unsupported files: {", ".join(skipped_files)}. Please upload PNG, JPG, JPEG, or WEBP only.', 'warning')
    
    # 61-80: Append to Excel file using openpyxl
    if processed_data:
        if save_to_excel(processed_data):
            flash(f'Successfully processed {len(processed_data)} visiting cards and saved to output.xlsx', 'success')
        else:
            flash('Error saving data to Excel file', 'error')
    else:
        if not skipped_files:  # Only show this if no files were skipped due to format
            flash('No valid data extracted from any uploaded files', 'error')
    
    # Clean up temporary files
    cleanup_temp_files(saved_file_paths)
    print("🧹 Cleanup completed")
    
    return redirect(url_for('main.index'))

@main_bp.route('/results')
def view_results():
    """
    101-120: View extracted results from Excel file
    """
    try:
        from openpyxl import load_workbook
        results = []
        
        # Get absolute path for Excel file
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        excel_path = os.path.join(project_root, 'output.xlsx')
        
        print(f"📊 Looking for Excel file at: {excel_path}")
        
        # Read Excel file if it exists
        if os.path.exists(excel_path):
            try:
                workbook = load_workbook(excel_path)
                worksheet = workbook.active
                
                # Get headers from first row
                headers = []
                for cell in worksheet[1]:
                    if cell.value:
                        headers.append(cell.value.lower())
                
                print(f"📋 Excel headers: {headers}")
                
                # Read data rows (skip header row)
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    if any(row):  # Skip empty rows
                        row_data = {}
                        for i, value in enumerate(row):
                            if i < len(headers):
                                row_data[headers[i]] = value or ''
                        
                        # Ensure all required fields exist
                        row_data.setdefault('name', '')
                        row_data.setdefault('email', '')
                        row_data.setdefault('phone', '')
                        row_data.setdefault('company', '')
                        row_data.setdefault('filename', '')
                        
                        results.append(row_data)
                
                workbook.close()
                print(f"✅ Loaded {len(results)} records from Excel")
                
            except Exception as excel_error:
                print(f"❌ Error reading Excel file: {excel_error}")
                flash('Error reading Excel file', 'error')
        else:
            print("📄 No Excel file found, displaying empty results")
        
        # Render results page with data
        return render_template('results.html', results=results)
        
    except Exception as e:
        print(f"❌ Error in view_results: {str(e)}")
        flash(f'Error reading results: {str(e)}', 'error')
        return redirect(url_for('main.index'))

@main_bp.route('/api/upload', methods=['POST'])
def api_upload():
    """
    121-150: API endpoint for file upload (JSON response)
    """
    try:
        # Similar processing as upload_files but returns JSON
        if 'files' not in request.files:
            return jsonify({'error': 'No files provided'}), 400
        
        uploaded_files = request.files.getlist('files')
        if not uploaded_files or all(file.filename == '' for file in uploaded_files):
            return jsonify({'error': 'No valid files provided'}), 400
        
        processed_data = []
        upload_folder = current_app.config['UPLOAD_FOLDER']
        os.makedirs(upload_folder, exist_ok=True)
        
        for file in uploaded_files:
            if file and file.filename != '':
                file_path = save_uploaded_file(file, upload_folder)
                if file_path:
                    raw_text = extract_text_from_image(file_path)
                    if raw_text:
                        structured_data = extract_data(raw_text)
                        if validate_extracted_data(structured_data):
                            structured_data['filename'] = file.filename
                            processed_data.append(structured_data)
                    # Clean up individual file
                    cleanup_temp_files([file_path])
        
        if processed_data:
            save_to_excel(processed_data)
            return jsonify({
                'success': True,
                'message': f'Processed {len(processed_data)} files and saved to output.xlsx',
                'data': processed_data
            })
        else:
            return jsonify({'error': 'No valid data extracted'}), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500
