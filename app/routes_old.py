# 1-10: Import modules
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app
import os
from app.ocr import extract_data_from_image_gemini  # Gemini Vision API extraction
from app.utils import save_uploaded_file, save_to_excel, cleanup_temp_files, allowed_file, get_recent_extractions  # Import utility functions

# 11-20: Blueprint creation
main_bp = Blueprint('main', __name__)  # Create main blueprint for routes

@main_bp.route('/')
def index():
    """
    11-20: Home page route - render file upload form with recent extractions
    """
    # Get recent extractions for display
    recent_extractions = get_recent_extractions(limit=5)
    
    # Render the main upload page with recent data
    return render_template('index.html', recent_extractions=recent_extractions)

@main_bp.route('/upload', methods=['POST'])
def upload_files():
    """
    21-100: Handle bulk file upload and Gemini Vision API extraction with progress tracking
    """
    print("🚀 Starting bulk file upload process...")
    
    # 11-20: Validate and read uploaded files
    if 'files' not in request.files:
        flash('No files selected', 'error')
        return redirect(url_for('main.index'))
    
    uploaded_files = request.files.getlist('files')
    
    if not uploaded_files or all(file.filename == '' for file in uploaded_files):
        flash('No files selected', 'error')
        return redirect(url_for('main.index'))
    
    # Initialize processing variables
    processed_data = []
    saved_file_paths = []
    upload_folder = current_app.config['UPLOAD_FOLDER']
    skipped_files = []
    total_files = len(uploaded_files)
    processed_count = 0
    
    os.makedirs(upload_folder, exist_ok=True)
    print(f"📁 Upload folder: {upload_folder}")
    print(f"📊 Total files to process: {total_files}")
    
    # 21-50: Handle multiple file uploads, validation, and OCR loop
    for idx, file in enumerate(uploaded_files):
        if file and file.filename != '':
            print(f"\n📄 Processing file {idx + 1}/{total_files}: {file.filename}")
            
            # Check file format before processing
            if not allowed_file(file.filename):
                print(f"❌ Unsupported format: {file.filename}")
                skipped_files.append(file.filename)
                continue
            
            try:
                # Save uploaded file with secure filename
                file_path = save_uploaded_file(file, upload_folder)
                
                if file_path:
                    saved_file_paths.append(file_path)
                    print(f"💾 File saved: {file_path}")
                    
                    # Read image as bytes for Gemini
                    with open(file_path, 'rb') as imgf:
                        image_bytes = imgf.read()
                    
                    # Extract structured data using Gemini
                    structured_data = extract_data_from_image_gemini(image_bytes)
                    structured_data['filename'] = file.filename
                    processed_data.append(structured_data)
                    processed_count += 1
                    print(f"✅ Extracted: {structured_data}")
                else:
                    print(f"❌ Failed to save {file.filename}")
                    
            except Exception as e:
                print(f"❌ Error processing {file.filename}: {str(e)}")
                continue
    
    # Handle results and messages
    if skipped_files:
        flash(f'Skipped {len(skipped_files)} unsupported files: {", ".join(skipped_files[:5])}{"..." if len(skipped_files) > 5 else ""}. Please upload PNG, JPG, JPEG, or WEBP only.', 'warning')
    
    # 51-70: Append to Excel using openpyxl
    if processed_data:
        if save_to_excel(processed_data):
            success_msg = f'Successfully processed {len(processed_data)} visiting cards and saved to output.xlsx'
            if skipped_files:
                success_msg += f' ({len(skipped_files)} files skipped)'
            flash(success_msg, 'success')
        else:
            flash('Error saving data to Excel file', 'error')
    else:
        if not skipped_files:  # Only show this if no files were skipped due to format
            flash('No valid data extracted from any uploaded files', 'error')
        else:
            flash('All uploaded files were skipped due to unsupported formats', 'warning')
    
    # Clean up temporary files
    cleanup_temp_files(saved_file_paths)
    print("🧹 Cleanup completed")
    print(f"📊 Final stats: {processed_count} processed, {len(skipped_files)} skipped, {total_files} total")
    
    return redirect(url_for('main.index'))

@main_bp.route('/results')
def view_results():
    """
    101-120: View extracted results from Excel file
    """
    try:
        from openpyxl import load_workbook
        results = []
        
        # Get absolute path for Excel file (from static/results/)
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        excel_path = os.path.join(project_root, 'static', 'results', 'output.xlsx')
        
        print(f"📊 Looking for Excel file at: {excel_path}")
        
        # Read Excel file if it exists
        if os.path.exists(excel_path):
            try:
                workbook = load_workbook(excel_path)
                worksheet = workbook.active
                
                # Get headers from first row
                headers = []
                for cell in worksheet[1]:
                    if cell.value:
                        headers.append(cell.value.lower())
                
                print(f"📋 Excel headers: {headers}")
                
                # Read data rows (skip header row)
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    if any(row):  # Skip empty rows
                        row_data = {}
                        for i, value in enumerate(row):
                            if i < len(headers):
                                row_data[headers[i]] = value or ''
                        
                        # Ensure all required fields exist
                        row_data.setdefault('name', '')
                        row_data.setdefault('email', '')
                        row_data.setdefault('phone', '')
                        row_data.setdefault('company', '')
                        row_data.setdefault('filename', '')
                        
                        results.append(row_data)
                
                workbook.close()
                print(f"✅ Loaded {len(results)} records from Excel")
                
            except Exception as excel_error:
                print(f"❌ Error reading Excel file: {excel_error}")
                flash('Error reading Excel file', 'error')
        else:
            print("📄 No Excel file found, displaying empty results")
        
        # Render results page with data
        return render_template('results.html', results=results)
        
    except Exception as e:
        print(f"❌ Error in view_results: {str(e)}")
        flash(f'Error reading results: {str(e)}', 'error')
        return redirect(url_for('main.index'))

@main_bp.route('/api/upload', methods=['POST'])
def api_upload():
    """
    121-150: API endpoint for file upload (JSON response)
    """
    try:
        # Similar processing as upload_files but returns JSON
        if 'files' not in request.files:
            return jsonify({'error': 'No files provided'}), 400
        
        uploaded_files = request.files.getlist('files')
        if not uploaded_files or all(file.filename == '' for file in uploaded_files):
            return jsonify({'error': 'No valid files provided'}), 400
        
        processed_data = []
        upload_folder = current_app.config['UPLOAD_FOLDER']
        os.makedirs(upload_folder, exist_ok=True)
        
        for file in uploaded_files:
            if file and file.filename != '':
                file_path = save_uploaded_file(file, upload_folder)
                if file_path:
                    # Read image as bytes for Gemini
                    with open(file_path, 'rb') as imgf:
                        image_bytes = imgf.read()
                    
                    # Extract structured data using Gemini
                    structured_data = extract_data_from_image_gemini(image_bytes)
                    structured_data['filename'] = file.filename
                    processed_data.append(structured_data)
                    # Clean up individual file
                    cleanup_temp_files([file_path])
        
        if processed_data:
            save_to_excel(processed_data)
            return jsonify({
                'success': True,
                'message': f'Processed {len(processed_data)} files and saved to output.xlsx',
                'data': processed_data
            })
        else:
            return jsonify({'error': 'No valid data extracted'}), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main_bp.route('/download')
def download_excel():
    """
    91-100: Route to download the Excel file
    """
    try:
        # Get absolute path for Excel file (from static/results/)
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        excel_path = os.path.join(project_root, 'static', 'results', 'output.xlsx')
        
        print(f"📥 Download request for Excel file: {excel_path}")
        
        # Check if Excel file exists
        if os.path.exists(excel_path):
            from flask import send_file
            print("✅ Sending Excel file to user")
            return send_file(
                excel_path, 
                as_attachment=True, 
                download_name='visiting_cards_data.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            print("❌ Excel file not found")
            flash('No data available for download. Please process some cards first.', 'warning')
            return redirect(url_for('main.index'))
            
    except Exception as e:
        print(f"❌ Error downloading Excel file: {str(e)}")
        flash(f'Error downloading file: {str(e)}', 'error')
        return redirect(url_for('main.index'))

@main_bp.route('/api/progress/<session_id>')
def get_progress(session_id):
    """
    71-90: Track progress count and send to frontend (for future implementation)
    """
    # This can be used for real-time progress tracking
    # For now, returning basic response
    return jsonify({'progress': 100, 'status': 'completed'})

@main_bp.route('/api/recent')
def api_recent_extractions():
    """
    111-130: API endpoint to get recent extractions for AJAX updates
    """
    try:
        recent_extractions = get_recent_extractions(limit=5)
        return jsonify({
            'success': True,
            'data': recent_extractions,
            'count': len(recent_extractions)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
