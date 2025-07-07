# 1-10: Importing modules
import os  # OS module for file operations
from openpyxl import Workbook  # Excel file handling
from openpyxl.chart import BarChart, PieChart, Reference, Series  # Excel charting
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # Excel styling
from werkzeug.utils import secure_filename  # Secure filename utility
import uuid  # UUID for generating unique filenames
from io import BytesIO  # For in-memory file handling
from app.mongo import load_extraction_data, get_all_labels  # Import MongoDB data loading functions
from collections import Counter  # For statistics
from datetime import datetime  # For date handling
import re  # For regex operations

# 11-20: File validation configuration
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp'}  # Supported image formats only

def allowed_file(filename):
    """
    21-30: Check if uploaded file has allowed extension
    """
    # Check if filename contains dot and extension is in allowed list
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_uploaded_file(file, upload_folder):
    """
    31-50: Save uploaded file with secure filename
    """
    if file and allowed_file(file.filename):
        # Generate unique filename to avoid conflicts
        file_extension = file.filename.rsplit('.', 1)[1].lower()
        unique_filename = f"{uuid.uuid4().hex}.{file_extension}"
        
        # Create secure filename
        filename = secure_filename(unique_filename)
        
        # Create full file path
        file_path = os.path.join(upload_folder, filename)
        
        # Save file to upload folder
        file.save(file_path)
        
        return file_path  # Return saved file path
    
    return None  # Return None if file is invalid

def generate_excel_from_mongo():
    """
    51-80: Generate Excel file in memory from MongoDB data (no local file storage)
    Returns BytesIO object containing Excel data for direct download
    """
    print("� Generating Excel file from MongoDB data...")
    
    try:
        # Load all data from MongoDB
        data_list = load_extraction_data()
        
        if not data_list:
            print("⚠️ No data found in MongoDB")
            return None
        
        # Create new workbook in memory
        workbook = Workbook()
        worksheet = workbook.active
        
        # Add comprehensive headers
        headers = ['Name', 'Phone', 'Email', 'Company', 'Website', 'Address', 'Filename', 'Timestamp', 
                  'Event Name', 'Event Description', 'Event Host', 'Event Date', 'Event Location']
        worksheet.append(headers)
        print("✅ Headers added to Excel file")
        
        # Add data rows
        rows_added = 0
        for data in data_list:
            row_data = [
                data.get('name', ''),
                data.get('phone', ''),
                data.get('email', ''),
                data.get('company', ''),
                data.get('website', ''),
                data.get('address', ''),
                data.get('filename', ''),
                data.get('timestamp', ''),
                # Event-related fields
                data.get('event_name', ''),
                data.get('event_description', ''),
                data.get('event_host', ''),
                data.get('event_date', ''),
                data.get('event_location', '')
            ]
            worksheet.append(row_data)
            rows_added += 1
        
        # Save to BytesIO object (in memory)
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        workbook.close()
        
        # Reset buffer position to beginning
        excel_buffer.seek(0)
        
        print(f"💾 Successfully generated Excel with {rows_added} rows in memory")
        return excel_buffer
        
    except Exception as e:
        print(f"❌ Error generating Excel: {str(e)}")
        return None

def cleanup_temp_files(file_paths):
    """
    81-100: Clean up temporary uploaded files
    """
    for file_path in file_paths:
        try:
            # Remove file if it exists
            if os.path.exists(file_path):
                os.remove(file_path)
                print(f"🧹 Cleaned up temporary file: {file_path}")
        except Exception as e:
            print(f"❌ Error cleaning up file {file_path}: {str(e)}")

def validate_extracted_data(data):
    """
    101-120: Validate extracted data before saving to MongoDB
    """
    # Check if at least name or email or phone is present
    required_fields = ['name', 'email', 'phone']
    has_required_data = any(data.get(field, '').strip() for field in required_fields)
    
    return has_required_data  # Return True if valid data found

def generate_advanced_analytics_report():
    """
    Generate comprehensive analytics report with multiple sheets and charts
    """
    print("📊 Generating advanced analytics report...")
    
    try:
        # Load all data from MongoDB
        data_list = load_extraction_data()
        
        if not data_list:
            print("⚠️ No data found in MongoDB")
            return None
        
        # Create new workbook
        workbook = Workbook()
        
        # Remove default sheet and create custom sheets
        workbook.remove(workbook.active)
        
        # Sheet 1: Complete Data
        data_sheet = workbook.create_sheet("Complete Data")
        _add_complete_data_sheet(data_sheet, data_list)
        
        # Sheet 2: Analytics Summary
        summary_sheet = workbook.create_sheet("Analytics Summary")
        _add_analytics_summary_sheet(summary_sheet, data_list)
        
        # Sheet 3: Company Analysis
        company_sheet = workbook.create_sheet("Company Analysis")
        _add_company_analysis_sheet(company_sheet, data_list)
        
        # Sheet 4: Geographic Analysis
        geo_sheet = workbook.create_sheet("Geographic Analysis")
        _add_geographic_analysis_sheet(geo_sheet, data_list)
        
        # Sheet 5: Contact Methods Analysis
        contact_sheet = workbook.create_sheet("Contact Analysis")
        _add_contact_analysis_sheet(contact_sheet, data_list)
        
        # Save to BytesIO object
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        workbook.close()
        excel_buffer.seek(0)
        
        print(f"💾 Successfully generated advanced analytics report")
        return excel_buffer
        
    except Exception as e:
        print(f"❌ Error generating advanced analytics report: {str(e)}")
        return None

def generate_filtered_excel_by_labels(label_ids, include_unlabeled=False):
    """
    Generate Excel file filtered by specific labels
    """
    print(f"🏷️ Generating Excel filtered by labels: {label_ids}")
    
    try:
        # Load all data from MongoDB
        data_list = load_extraction_data()
        
        if not data_list:
            print("⚠️ No data found in MongoDB")
            return None
        
        # Filter data by labels
        filtered_data = []
        for data in data_list:
            # Check current label assignment
            data_label_id = data.get('label_id')
            
            # Check if data matches selected labels
            has_selected_label = data_label_id in label_ids if data_label_id else False
            
            # Check if data is unlabeled and should be included
            is_unlabeled = not data_label_id or data_label_id == ''
            
            if has_selected_label or (include_unlabeled and is_unlabeled):
                filtered_data.append(data)
        
        if not filtered_data:
            print("⚠️ No data matches the selected labels")
            return None
        
        # Create workbook with filtered data
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Filtered by Labels"
        
        # Add headers with label information
        headers = ['Name', 'Phone', 'Email', 'Company', 'Website', 'Address', 'Label', 
                  'Filename', 'Timestamp', 'Event Name', 'Event Description', 'Event Host', 'Event Date', 'Event Location']
        worksheet.append(headers)
        
        # Add filtered data rows
        for data in filtered_data:
            # Format labels for display
            label_display = data.get('label_name', 'Unlabeled')
            
            row_data = [
                data.get('name', ''),
                data.get('phone', ''),
                data.get('email', ''),
                data.get('company', ''),
                data.get('website', ''),
                data.get('address', ''),
                label_display,
                data.get('filename', ''),
                data.get('timestamp', ''),
                data.get('event_name', ''),
                data.get('event_description', ''),
                data.get('event_host', ''),
                data.get('event_date', ''),
                data.get('event_location', '')
            ]
            worksheet.append(row_data)
        
        # Save to BytesIO object
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        workbook.close()
        excel_buffer.seek(0)
        
        print(f"💾 Successfully generated filtered Excel with {len(filtered_data)} rows")
        return excel_buffer
        
    except Exception as e:
        print(f"❌ Error generating filtered Excel by labels: {str(e)}")
        return None

def generate_filtered_excel_by_countries(countries):
    """
    Generate Excel file filtered by specific countries
    """
    print(f"🌍 Generating Excel filtered by countries: {countries}")
    
    try:
        # Load all data from MongoDB
        data_list = load_extraction_data()
        
        if not data_list:
            print("⚠️ No data found in MongoDB")
            return None
        
        # Filter data by countries
        filtered_data = []
        for data in data_list:
            data_country = data.get('country', '')
            if data_country in countries:
                filtered_data.append(data)
        
        if not filtered_data:
            print("⚠️ No data matches the selected countries")
            return None
        
        # Create workbook with filtered data
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Filtered by Countries"
        
        # Add headers with country information
        headers = ['Name', 'Phone', 'Email', 'Company', 'Website', 'Address', 'Country', 'Flag',
                  'Filename', 'Timestamp', 'Event Name', 'Event Description', 'Event Host', 'Event Date', 'Event Location']
        worksheet.append(headers)
        
        # Add filtered data rows
        for data in filtered_data:
            row_data = [
                data.get('name', ''),
                data.get('phone', ''),
                data.get('email', ''),
                data.get('company', ''),
                data.get('website', ''),
                data.get('address', ''),
                data.get('country', ''),
                data.get('flag', ''),
                data.get('filename', ''),
                data.get('timestamp', ''),
                data.get('event_name', ''),
                data.get('event_description', ''),
                data.get('event_host', ''),
                data.get('event_date', ''),
                data.get('event_location', '')
            ]
            worksheet.append(row_data)
        
        # Save to BytesIO object
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        workbook.close()
        excel_buffer.seek(0)
        
        print(f"💾 Successfully generated filtered Excel with {len(filtered_data)} rows")
        return excel_buffer
        
    except Exception as e:
        print(f"❌ Error generating filtered Excel by countries: {str(e)}")
        return None

# Helper functions for advanced analytics report
def _add_complete_data_sheet(worksheet, data_list):
    """Add complete data to the first sheet with formatting"""
    # Add headers
    headers = ['Name', 'Phone', 'Email', 'Company', 'Website', 'Address', 'Country', 'Label',
              'Filename', 'Timestamp', 'Event Name', 'Event Description', 'Event Host', 'Event Date', 'Event Location']
    worksheet.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Add data rows
    for data in data_list:
        label_display = data.get('label_name', 'Unlabeled')
        
        row_data = [
            data.get('name', ''),
            data.get('phone', ''),
            data.get('email', ''),
            data.get('company', ''),
            data.get('website', ''),
            data.get('address', ''),
            data.get('country', ''),
            label_display,
            data.get('filename', ''),
            data.get('timestamp', ''),
            data.get('event_name', ''),
            data.get('event_description', ''),
            data.get('event_host', ''),
            data.get('event_date', ''),
            data.get('event_location', '')
        ]
        worksheet.append(row_data)

def _add_analytics_summary_sheet(worksheet, data_list):
    """Add analytics summary with key statistics"""
    # Title
    worksheet['A1'] = "Analytics Summary"
    worksheet['A1'].font = Font(size=16, bold=True)
    
    # Basic statistics
    total_cards = len(data_list)
    cards_with_email = sum(1 for data in data_list if data.get('email', '').strip())
    cards_with_phone = sum(1 for data in data_list if data.get('phone', '').strip())
    cards_with_company = sum(1 for data in data_list if data.get('company', '').strip())
    
    stats = [
        ["Metric", "Value"],
        ["Total Cards Processed", total_cards],
        ["Cards with Email", cards_with_email],
        ["Cards with Phone", cards_with_phone],
        ["Cards with Company", cards_with_company],
        ["Email Coverage", f"{(cards_with_email/total_cards)*100:.1f}%" if total_cards > 0 else "0%"],
        ["Phone Coverage", f"{(cards_with_phone/total_cards)*100:.1f}%" if total_cards > 0 else "0%"],
        ["Company Coverage", f"{(cards_with_company/total_cards)*100:.1f}%" if total_cards > 0 else "0%"]
    ]
    
    for i, row in enumerate(stats, start=3):
        worksheet[f'A{i}'] = row[0]
        worksheet[f'B{i}'] = row[1]
        if i == 3:  # Header row
            worksheet[f'A{i}'].font = Font(bold=True)
            worksheet[f'B{i}'].font = Font(bold=True)

def _add_company_analysis_sheet(worksheet, data_list):
    """Add company analysis with top companies"""
    worksheet['A1'] = "Company Analysis"
    worksheet['A1'].font = Font(size=16, bold=True)
    
    # Count companies
    companies = [data.get('company', '').strip() for data in data_list if data.get('company', '').strip()]
    company_counts = Counter(companies)
    
    # Top companies
    worksheet['A3'] = "Top Companies"
    worksheet['A3'].font = Font(bold=True)
    worksheet['A4'] = "Company"
    worksheet['B4'] = "Count"
    
    for i, (company, count) in enumerate(company_counts.most_common(20), start=5):
        worksheet[f'A{i}'] = company
        worksheet[f'B{i}'] = count

def _add_geographic_analysis_sheet(worksheet, data_list):
    """Add geographic analysis by countries"""
    worksheet['A1'] = "Geographic Analysis"
    worksheet['A1'].font = Font(size=16, bold=True)
    
    # Count countries
    countries = [data.get('country', '').strip() for data in data_list if data.get('country', '').strip()]
    country_counts = Counter(countries)
    
    # Top countries
    worksheet['A3'] = "Countries Distribution"
    worksheet['A3'].font = Font(bold=True)
    worksheet['A4'] = "Country"
    worksheet['B4'] = "Count"
    worksheet['C4'] = "Percentage"
    
    total_with_country = len(countries)
    for i, (country, count) in enumerate(country_counts.most_common(), start=5):
        worksheet[f'A{i}'] = country
        worksheet[f'B{i}'] = count
        worksheet[f'C{i}'] = f"{(count/total_with_country)*100:.1f}%" if total_with_country > 0 else "0%"

def _add_contact_analysis_sheet(worksheet, data_list):
    """Add contact methods analysis"""
    worksheet['A1'] = "Contact Methods Analysis"
    worksheet['A1'].font = Font(size=16, bold=True)
    
    # Analyze contact methods
    total_cards = len(data_list)
    email_only = sum(1 for data in data_list if data.get('email', '').strip() and not data.get('phone', '').strip())
    phone_only = sum(1 for data in data_list if data.get('phone', '').strip() and not data.get('email', '').strip())
    both_contacts = sum(1 for data in data_list if data.get('email', '').strip() and data.get('phone', '').strip())
    no_contact = sum(1 for data in data_list if not data.get('email', '').strip() and not data.get('phone', '').strip())
    
    contact_stats = [
        ["Contact Method", "Count", "Percentage"],
        ["Email Only", email_only, f"{(email_only/total_cards)*100:.1f}%" if total_cards > 0 else "0%"],
        ["Phone Only", phone_only, f"{(phone_only/total_cards)*100:.1f}%" if total_cards > 0 else "0%"],
        ["Both Email & Phone", both_contacts, f"{(both_contacts/total_cards)*100:.1f}%" if total_cards > 0 else "0%"],
        ["No Contact Info", no_contact, f"{(no_contact/total_cards)*100:.1f}%" if total_cards > 0 else "0%"]
    ]
    
    for i, row in enumerate(contact_stats, start=3):
        worksheet[f'A{i}'] = row[0]
        worksheet[f'B{i}'] = row[1]
        worksheet[f'C{i}'] = row[2]
        if i == 3:  # Header row
            for col in ['A', 'B', 'C']:
                worksheet[f'{col}{i}'].font = Font(bold=True)
